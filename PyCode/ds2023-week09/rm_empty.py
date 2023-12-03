import xlrd
import openpyxl

def remove_leading_spaces(input_file, output_file):
    try:
        print(f"处理文件: {input_file}")

        xls_workbook = xlrd.open_workbook(input_file)
        wb = openpyxl.Workbook()

        # 遍历每个表单
        for sheet_index in range(xls_workbook.nsheets):
            sheet = xls_workbook.sheet_by_index(sheet_index)
            ws = wb.create_sheet(title=sheet.name)
            # 遍历每个单元格
            for row_index in range(sheet.nrows):
                for col_index in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_index, col_index)
                    if isinstance(cell_value, str):
                        cell_value = cell_value.strip()
                    ws.cell(row=row_index + 1, column=col_index + 1, value=cell_value)

        # 删除默认创建的空白表单
        del wb["Sheet"]

        wb.save(output_file)
        print(f"处理完成，已保存到 {output_file}")
    except Exception as e:
        print(f"发生错误: {e}")

if __name__ == "__main__":
    input_file = "C:\\Users\\19657\\Documents\\DS2023-pre\\2018-2023\\gpu.xls"
    output_file = "C:\\Users\\19657\\Documents\\DS2023-pre\\2018-2023\\gpu_modified.xlsx"
    remove_leading_spaces(input_file, output_file)